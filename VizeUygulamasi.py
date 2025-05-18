import sys
import re
import sqlite3
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget, QVBoxLayout, QFormLayout, QTextEdit,
    QLabel, QLineEdit, QComboBox, QRadioButton, QGroupBox, QPushButton,
    QCheckBox, QDateEdit, QMessageBox, QTableWidget, QTableWidgetItem, QHBoxLayout
)
from PyQt5.QtCore import QDate
from openpyxl import Workbook, load_workbook
import os
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap, QIcon



def create_connection():
    conn = sqlite3.connect(r"C:\Users\lenovo\OneDrive\Masaüstü\Randevular\randevular.db")
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS randevular (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        adsoyad TEXT,
        pasaport TEXT,
        email TEXT,
        randevu_tarihi TEXT,
        randevu_saati TEXT,
        ofis TEXT
    )
    """)
    conn.commit()
    return conn


class RandevuUygulamasi(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Vize Randevu Sistemi")
        self.setGeometry(300, 100, 800, 600)

        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        # Sekmelerin tanımlanması
        self.tab0 = QWidget()
        self.tab1 = QWidget()
        self.tab2 = QWidget()
        self.tab3 = QWidget()
        self.tab4 = QWidget()
        self.tab5 = QWidget()
        self.tab6 = QWidget()

        self.tabs.addTab(self.tab0,"Giriş")
        self.tabs.addTab(self.tab1, "Kişisel Bilgiler")
        self.tabs.addTab(self.tab2, "Ülke Seçimleri")
        self.tabs.addTab(self.tab3, "Randevu Detayları")
        self.tabs.addTab(self.tab4, "Ek İşlemler")
        self.tabs.addTab(self.tab5, "Randevuyu Onayla")
        self.tabs.addTab(self.tab6, "Geçmiş Randevular")

        self.init_tab0()
        self.init_tab1()
        self.init_tab2()
        self.init_tab3()
        self.init_tab4()
        self.init_tab5()
        self.init_tab6()
        self.tabs.currentChanged.connect(self.handle_tab_change)


        
        self.tabs.setCurrentIndex(0)
        self.veriler = {}

    def init_tab0(self):
        layout = QVBoxLayout()

        welcome_label = QLabel("🌍 Vize Randevu Sistemine Hoş Geldiniz")
        welcome_label.setAlignment(Qt.AlignCenter)
        welcome_label.setStyleSheet("font-size: 24px; font-weight: bold;")

        desc_label = QLabel("Devam etmek için aşağıdaki butona tıklayın.")
        desc_label.setAlignment(Qt.AlignCenter)
        desc_label.setStyleSheet("font-size: 14px;")

        logo = QLabel()
        pixmap = QPixmap("logo.png").scaled(150, 150, Qt.KeepAspectRatio)
        logo.setPixmap(pixmap)
        logo.setAlignment(Qt.AlignCenter)

        self.start_button = QPushButton("Randevuya Başla")
        self.start_button.clicked.connect(self.handle_start_button)  # Burayı ekle
        self.start_button.setFixedHeight(40)
        self.start_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 16px;")

        layout.addStretch()
        layout.addWidget(welcome_label)
        layout.addWidget(desc_label)
        layout.addWidget(logo)
        layout.addWidget(self.start_button, alignment=Qt.AlignCenter)
        layout.addStretch()

        self.tab0.setLayout(layout)


    def init_tab1(self):
        main_layout = QVBoxLayout()

        # --- Kişisel Bilgiler Grubu ---
        personal_group = QGroupBox("Kişisel Bilgiler")
        personal_layout = QFormLayout()

        self.name_input = QLineEdit()
        self.passport_input = QLineEdit()
        self.tc_input = QLineEdit()
        self.birth_date_input = QDateEdit()
        self.birth_date_input.setCalendarPopup(True)
        self.birth_date_input.setDate(QDate(2000, 1, 1))

        self.gender_male = QRadioButton("Erkek")
        self.gender_female = QRadioButton("Kadın")
        gender_layout = QHBoxLayout()
        gender_layout.addWidget(self.gender_male)
        gender_layout.addWidget(self.gender_female)

        personal_layout.addRow("Ad Soyad:", self.name_input)
        personal_layout.addRow("Pasaport No:", self.passport_input)
        personal_layout.addRow("T.C. Kimlik No:", self.tc_input)
        personal_layout.addRow("Doğum Tarihi:", self.birth_date_input)
        personal_layout.addRow("Cinsiyet:", gender_layout)

        personal_group.setLayout(personal_layout)

        # --- İletişim Bilgileri Grubu ---
        contact_group = QGroupBox("İletişim Bilgileri")
        contact_layout = QFormLayout()

        self.email_input = QLineEdit()
        self.phone_input = QLineEdit()
        self.address_input = QLineEdit()

        contact_layout.addRow("E-posta:", self.email_input)
        contact_layout.addRow("Telefon (örnek: +90 123 456 7890):", self.phone_input)
        contact_layout.addRow("Adres:", self.address_input)

        contact_group.setLayout(contact_layout)

        # --- Diğer Bilgiler Grubu ---
        extra_group = QGroupBox("Ek Bilgiler")
        extra_layout = QFormLayout()

        self.nationality_combo = QComboBox()
        self.nationality_combo.addItems(["Türkiye", "Almanya", "Fransa", "İngiltere", "Amerika", "Diğer"])

        self.secret_question_combo = QComboBox()
        self.secret_question_combo.addItems([
            "İlk evcil hayvanınızın adı?",
            "İlkokul öğretmeninizin adı?",
            "Doğduğunuz şehir?",
        ])
        self.secret_answer_input = QLineEdit()

        extra_layout.addRow("Uyruk:", self.nationality_combo)
        extra_layout.addRow("Gizli Soru:", self.secret_question_combo)
        extra_layout.addRow("Cevap:", self.secret_answer_input)

        extra_group.setLayout(extra_layout)

        # --- Onay Butonu ---
        self.validate_button = QPushButton("Devam Et")
        self.validate_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 16px;")
        self.validate_button.clicked.connect(self.validate_tab1)

        # --- Ana Layout'a Ekle ---
        main_layout.addWidget(personal_group)
        main_layout.addWidget(contact_group)
        main_layout.addWidget(extra_group)
        main_layout.addWidget(self.validate_button)
        self.tab1.setLayout(main_layout)

    def init_tab2(self):
        layout = QVBoxLayout()

        # --- Yaşadığınız Ülke ve Şehir ---
        location_group = QGroupBox("İkamet Bilgileri")
        location_layout = QFormLayout()
        self.country_combo = QComboBox()
        self.country_combo.addItems(["Türkiye", "Almanya", "Amerika", "Fransa", "İngiltere", "Diğer"])
        self.city_input = QLineEdit()
        location_layout.addRow("Yaşadığınız Ülke:", self.country_combo)
        location_layout.addRow("Şehir (isteğe bağlı):", self.city_input)
        location_group.setLayout(location_layout)

        # --- Vize Bilgileri ---
        visa_group = QGroupBox("Vize Bilgileri")
        visa_layout = QFormLayout()
        self.visa_country_combo = QComboBox()
        self.visa_country_combo.addItems(["Almanya", "Fransa", "İngiltere", "Amerika", "Diğer"])

        self.visa_type_combo = QComboBox()
        self.visa_type_combo.addItems(["Turistik", "Eğitim", "Çalışma", "Aile Birleşimi", "Sağlık", "Diğer"])

        self.visa_reason_input = QTextEdit()
        self.visa_reason_input.setPlaceholderText("Lütfen neden vize almak istediğinizi açıklayın...")

        self.passport_type_combo = QComboBox()
        self.passport_type_combo.addItems(["Umuma Mahsus (Bordo)", "Hizmet (Gri)", "Hususi (Yeşil)", "Diplomatik (Siyah)"])

        visa_layout.addRow("Vize Başvurusu Yapmak İstediğiniz Ülke:", self.visa_country_combo)
        visa_layout.addRow("Vize Türü:", self.visa_type_combo)
        visa_layout.addRow("Pasaport Türü:", self.passport_type_combo)
        visa_layout.addRow("Vize Alma Gerekçesi:", self.visa_reason_input)
        visa_group.setLayout(visa_layout)

        # --- Daha Önce Gidildi Mi? ---
        history_group = QGroupBox("Seyahat Geçmişi")
        history_layout = QFormLayout()
        self.visited_before_combo = QComboBox()
        self.visited_before_combo.addItems(["Hayır", "Evet"])
        self.visited_explanation_input = QLineEdit()
        self.visited_explanation_input.setPlaceholderText("Eğer gittiyseniz yıl ve amaç belirtin...")
        history_layout.addRow("Daha önce bu ülkeye gittiniz mi?", self.visited_before_combo)
        history_layout.addRow("Açıklama (varsa):", self.visited_explanation_input)
        history_group.setLayout(history_layout)

        # --- Seyahat Süresi ---
        duration_group = QGroupBox("Seyahat Süresi")
        duration_layout = QFormLayout()
        self.travel_start_date = QDateEdit()
        self.travel_start_date.setCalendarPopup(True)
        self.travel_start_date.setDate(QDate.currentDate())

        self.travel_end_date = QDateEdit()
        self.travel_end_date.setCalendarPopup(True)
        self.travel_end_date.setDate(QDate.currentDate().addDays(7))

        duration_layout.addRow("Seyahatin Başlangıç Tarihi:", self.travel_start_date)
        duration_layout.addRow("Seyahatin Bitiş Tarihi:", self.travel_end_date)
        duration_group.setLayout(duration_layout)

        # --- Devam Et Butonu ---
        self.validate_tab2_button = QPushButton("Devam Et")
        self.validate_tab2_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 16px;")
        self.validate_tab2_button.clicked.connect(self.validate_tab2)

        # Ana Layout'a Ekle
        layout.addWidget(location_group)
        layout.addWidget(visa_group)
        layout.addWidget(history_group)
        layout.addWidget(duration_group)
        layout.addWidget(self.validate_tab2_button)

        self.tab2.setLayout(layout)

    def init_tab3(self):
        layout = QVBoxLayout()
        appointment_group = QGroupBox("Randevu Bilgileri")
        form_layout = QFormLayout()

        self.date_picker = QDateEdit()
        self.date_picker.setCalendarPopup(True)
        self.date_picker.setDate(QDate.currentDate())
        self.date_picker.setMinimumDate(QDate.currentDate())
        self.date_picker.setDisplayFormat("dd.MM.yyyy")
        self.date_picker.calendarWidget().setGridVisible(True)

        def disable_weekends(date):
            return date.dayOfWeek() not in (6, 7)

        def handle_date_click(date):
            if not disable_weekends(date):
                QMessageBox.warning(self, "Geçersiz Tarih", "Hafta sonu seçilemez!")
                self.date_picker.setDate(QDate.currentDate())

        self.date_picker.calendarWidget().clicked.connect(handle_date_click)

        def generate_half_hour_times():
            return [f"{h:02d}:{m:02d}" for h in range(9, 18) for m in (0, 30)]
 
        self.time_combo = QComboBox()
        self.time_combo.addItems(generate_half_hour_times())

        self.location_combo = QComboBox()
        self.location_combo.addItems(["İstanbul", "Ankara", "İzmir", "Bursa", "Antalya", "Diğer"])

        self.office_combo = QComboBox()
        self.office_combo.addItems(["VFS Global", "TLS Contact", "Konsolosluk", "Diğer"])

        self.appointment_type_combo = QComboBox()
        self.appointment_type_combo.addItems(["Yüz Yüze", "Online Görüşme", "Telefonla Görüşme"])

        self.delivery_combo = QComboBox()
        self.delivery_combo.addItems(["Elden Teslim", "Kargo ile Gönderim", "Online Yükleme (varsa sistem üzerinden)"])

        self.additional_docs_checkbox = QCheckBox("Ek belgeler isteniyor olabilir...")
        self.additional_notes_input = QTextEdit()
        self.additional_notes_input.setPlaceholderText("Özel bir notunuz varsa buraya yazabilirsiniz...")
        self.reminder_checkbox = QCheckBox("Randevu gününde bana hatırlatma gönderilsin")

        form_layout.addRow("Randevu Tarihi:", self.date_picker)
        form_layout.addRow("Randevu Saati:", self.time_combo)
        form_layout.addRow("Randevu Şehri:", self.location_combo)
        form_layout.addRow("Vize Merkezi / Ofis:", self.office_combo)
        form_layout.addRow("Randevu Türü:", self.appointment_type_combo)
        form_layout.addRow("Belge Teslim Şekli:", self.delivery_combo)
        form_layout.addRow("Ek Belgeler:", self.additional_docs_checkbox)
        form_layout.addRow("Ek Notlar:", self.additional_notes_input)
        form_layout.addRow("", self.reminder_checkbox)

        appointment_group.setLayout(form_layout)
        layout.addWidget(appointment_group)

        self.validate_tab3_button = QPushButton("Devam Et")
        self.validate_tab3_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 16px;")
        self.validate_tab3_button.clicked.connect(self.confirm_appointment)
        layout.addWidget(self.validate_tab3_button)

        self.tab3.setLayout(layout)
    
    
    def init_tab4(self):
        layout = QVBoxLayout()

        # --- Ek İşlemler ---
        extra_group = QGroupBox("Ek İşlemler")
        extra_layout = QVBoxLayout()

        # Sabit Ücret
        self.base_fee_label = QLabel("Sabit Randevu Ücreti: 50 USD")
        self.total_fee_label = QLabel("Toplam Ücret: 50 USD")

        # Ek hizmetler
        self.insurance_checkbox = QCheckBox("Seyahat Sigortası Yardımı (40 USD)")
        self.accommodation_checkbox = QCheckBox("Konaklama Yardımı (100 USD)")
        self.document_service_checkbox = QCheckBox("Belge Düzenleme Yardımı (50 USD)")
        self.translator_service_checkbox = QCheckBox("Tercüman Yardımı (60 USD)")
        self.fast_track_checkbox = QCheckBox("Hızlı İşlem Hizmeti (80 USD)")
        self.phone_email_support_checkbox = QCheckBox("Telefon ve E-posta Desteği (20 USD)")

        # Not alanı
        self.note_input = QTextEdit()
        self.note_input.setPlaceholderText("Ek açıklamalarınızı buraya yazabilirsiniz...")

        # Tahmini sonuçlanma süresi (şimdilik sabit)
        self.estimated_time_label = QLabel("Tahmini Sonuçlanma Süresi: 5 iş günü")

        # Layout ekleme
        extra_layout.addWidget(self.base_fee_label)
        extra_layout.addWidget(self.insurance_checkbox)
        extra_layout.addWidget(self.accommodation_checkbox)
        extra_layout.addWidget(self.document_service_checkbox)
        extra_layout.addWidget(self.translator_service_checkbox)
        extra_layout.addWidget(self.fast_track_checkbox)
        extra_layout.addWidget(self.phone_email_support_checkbox)
        extra_layout.addWidget(QLabel("Notlar:"))
        extra_layout.addWidget(self.note_input)
        extra_layout.addWidget(self.estimated_time_label)
        extra_layout.addWidget(self.total_fee_label)

        extra_group.setLayout(extra_layout)

        # Devam Et butonu
        self.proceed_button = QPushButton("Devam Et")
        self.proceed_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 16px;")
        self.proceed_button.clicked.connect(self.proceed_to_confirmation)

        layout.addWidget(extra_group)
        layout.addWidget(self.proceed_button)
        self.tab4.setLayout(layout)

        # Checkbox değişimlerine tepki vererek ücret hesapla
        self.insurance_checkbox.stateChanged.connect(self.update_total_fee)
        self.accommodation_checkbox.stateChanged.connect(self.update_total_fee)
        self.document_service_checkbox.stateChanged.connect(self.update_total_fee)
        self.translator_service_checkbox.stateChanged.connect(self.update_total_fee)
        self.fast_track_checkbox.stateChanged.connect(self.update_total_fee)
        self.phone_email_support_checkbox.stateChanged.connect(self.update_total_fee)


    def update_total_fee(self):
        base_fee = 50
        total_fee = base_fee
        selected_services = []

        if self.insurance_checkbox.isChecked():
            total_fee += 40
            selected_services.append("Seyahat Sigortası Yardımı")

        if self.accommodation_checkbox.isChecked():
            total_fee += 100
            selected_services.append("Konaklama Yardımı")

        if self.document_service_checkbox.isChecked():
            total_fee += 50
            selected_services.append("Belge Düzenleme Yardımı")

        if self.translator_service_checkbox.isChecked():
            total_fee += 60
            selected_services.append("Tercüman Yardımı")

        if self.fast_track_checkbox.isChecked():
            total_fee += 80
            selected_services.append("Hızlı İşlem Hizmeti")

        if self.phone_email_support_checkbox.isChecked():
            total_fee += 20
            selected_services.append("Telefon ve E-posta Desteği")

        # Toplam ücreti güncelle
        self.total_fee_label.setText(f"Toplam Ücret: {total_fee} USD")

        # Tahmini sonuç süresi - hızlı işlem varsa kısalt
        if self.fast_track_checkbox.isChecked():
           self.estimated_time_label.setText("Tahmini Sonuçlanma Süresi: 2 iş günü")
        else:
           self.estimated_time_label.setText("Tahmini Sonuçlanma Süresi: 5 iş günü")


    def init_tab5(self):
        layout = QVBoxLayout()

        self.summary_label = QLabel("Randevu bilgilerinizi kontrol edin ve onaylayın.")
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)

        layout.addWidget(self.summary_label)
        layout.addWidget(self.summary_text)

        self.confirm_button = QPushButton("Onayla ve Randevuyu Kaydet")
        self.confirm_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 16px;")
        self.confirm_button.clicked.connect(self.save_appointment)

        layout.addWidget(self.confirm_button)
        self.tab5.setLayout(layout)
    
    def write_to_excel(self, data):
        filename = "randevular.xlsx"
    
        if os.path.exists(filename):
           workbook = load_workbook(filename)
           sheet = workbook.active
        else:
           workbook = Workbook()
           sheet = workbook.active
           # Başlık satırı
           sheet.append(["Ad Soyad", "Pasaport", "Randevu Tarihi", "Randevu Saati", "Ofis"])
    
        # Verileri ekle
        sheet.append(data)
        workbook.save(filename)

    def init_tab6(self):
        layout = QVBoxLayout()

        # --- Geçmiş Randevular ---
        past_appointments_group = QGroupBox("Geçmiş Randevular")
        past_appointments_layout = QVBoxLayout()

        self.past_appointments_table = QTableWidget()
        self.past_appointments_table.setColumnCount(5)
        self.past_appointments_table.setHorizontalHeaderLabels(["Ad Soyad", "Pasaport No", "Randevu Tarihi", "Randevu Saati", "Ofis"])

        past_appointments_layout.addWidget(self.past_appointments_table)
        past_appointments_group.setLayout(past_appointments_layout)

        layout.addWidget(past_appointments_group)

        self.tab6.setLayout(layout)
        
        self.refresh_button = QPushButton("Yenile")
        self.refresh_button.setStyleSheet("background-color: #4CAF50; color: white; font-size: 16px;")
        self.refresh_button.clicked.connect(self.load_past_appointments)
        layout.addWidget(self.refresh_button)


    # Giriş Doğrulama Fonksiyonu
    def validate_tab1(self):
        name = self.name_input.text().strip()
        if not name or len(name.split()) < 2:
            QMessageBox.warning(self, "Geçersiz Giriş", "Ad ve Soyad en az iki kelime olmalı.")
            return

        passport = self.passport_input.text().strip()
        if not passport or len(passport) < 6 or len(passport) > 9:
            QMessageBox.warning(self, "Geçersiz Giriş", "Pasaport No boş olamaz ve 6 ile 9 haneli olmalıdır.")
            return

        tc = self.tc_input.text().strip()
        if not tc.isdigit() or len(tc) != 11:
            QMessageBox.warning(self, "Geçersiz Giriş", "T.C. Kimlik Numarası 11 haneli ve yalnızca rakamlardan oluşmalıdır.")
            return

        if not (self.gender_male.isChecked() or self.gender_female.isChecked()):
            QMessageBox.warning(self, "Geçersiz Giriş", "Lütfen bir cinsiyet seçin.")
            return
        
        email = self.email_input.text().strip()
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            QMessageBox.warning(self, "Geçersiz Giriş", "Geçerli bir e-posta adresi girin.")
            return
        
        phone = self.phone_input.text().strip()
        pattern = r"^\+\d{1,4}([ -]?\d{2,4}){2,4}$"
        if not re.match(pattern, phone):
            QMessageBox.warning(self, "Geçersiz Giriş", "Geçerli bir telefon numarası girin.")
            return

        address = self.address_input.text().strip()
        if len(address.split()) < 2:
            QMessageBox.warning(self, "Geçersiz Giriş", "Adres en az iki kelime olmalıdır.")
            return

        # Verileri Kaydetmeden Önce Doğrulama Başarıyla Geçti
        self.veriler['name'] = name
        self.veriler['passport'] = passport
        self.veriler['tc'] = tc
        self.veriler['email'] = email
        self.veriler['phone'] = phone
        self.veriler['address'] = address

        self.tabs.setCurrentIndex(1)  # 1. sekmeye geçiş (Ülke Seçimleri)
        QMessageBox.information(self, "Başarılı", "Girişler başarılı. Şimdi ülke seçimlerine geçebilirsiniz.")
        
        
        
    def validate_tab2(self):
        errors = []
        if not self.visa_country_combo.currentText():
            errors.append("Vize başvurusu yapılacak ülkeyi seçiniz.")

        if not self.visa_type_combo.currentText():
            errors.append("Vize türünü seçiniz.")

        if not self.visa_reason_input.toPlainText().strip():
            errors.append("Vize gerekçesini açıklamalısınız.")

        if self.travel_end_date.date() < self.travel_start_date.date():
            errors.append("Seyahat bitiş tarihi başlangıç tarihinden önce olamaz.")

        if errors:
            QMessageBox.warning(self, "Geçersiz Giriş", "\n".join(errors))
        else:
            self.tabs.setCurrentIndex(2)  # 2. sekmeye geçiş (Randevu Detayları)
            QMessageBox.information(self, "Başarılı", "Ülke seçimleri ve seyahat bilgileri başarılı!")

    def confirm_appointment(self):
        QMessageBox.information(self, "Randevu Detayları Onaylandı", "Randevu detayları başarıyla kaydedildi.\nEk işlemlere geçtiniz.")
        self.tabs.setCurrentIndex(3)

    def proceed_to_confirmation(self):
        self.tabs.setCurrentIndex(4)  # Tab 5'e geç

        summary = f"""
🧍 Ad Soyad: {self.name_input.text()}
🛂 Pasaport No: {self.passport_input.text()}
🆔 T.C. Kimlik No: {self.tc_input.text()}
📧 E-posta: {self.email_input.text()}
📞 Telefon: {self.phone_input.text()}
🏠 Adres: {self.address_input.text()}

🌍 Vize Ülkesi: {self.visa_country_combo.currentText()}
🎫 Vize Türü: {self.visa_type_combo.currentText()}
📄 Pasaport Türü: {self.passport_type_combo.currentText()}
📝 Gerekçe: {self.visa_reason_input.toPlainText()}

📅 Randevu Tarihi: {self.date_picker.date().toString("dd.MM.yyyy")}
⏰ Randevu Saati: {self.time_combo.currentText()}
🏢 Randevu Ofisi: {self.office_combo.currentText()}

💰 Toplam Ücret: {self.total_fee_label.text()}
    """.strip()

        self.summary_text.setPlainText(summary)

    def save_appointment(self):
        adsoyad = self.name_input.text()
        pasaport = self.passport_input.text()
        tarih = self.date_picker.date().toString("yyyy-MM-dd")
        saat = self.time_combo.currentText()
        ofis = self.office_combo.currentText()

        conn = create_connection()
        c = conn.cursor()
        c.execute("INSERT INTO randevular (adsoyad, pasaport, randevu_tarihi, randevu_saati, ofis) VALUES (?, ?, ?, ?, ?)",
              (adsoyad, pasaport, tarih, saat, ofis))
        conn.commit()
        conn.close()

        # Excel'e yaz
        excel_data = [adsoyad, pasaport, tarih, saat, ofis]
        self.write_to_excel(excel_data)

        QMessageBox.information(self, "Başarılı", "Randevu başarıyla kaydedildi.")
        self.tabs.setCurrentIndex(5)  # Geçmiş randevular sekmesine geç

        self.load_past_appointments()

    def handle_tab_change(self, index):
        if index == 5:  # 6. sekme index 5'tir (0'dan başladığı için)
           self.load_past_appointments()

    def handle_start_button(self):
        self.tabs.setCurrentIndex(1)  # Tab 1'e geç
 

    def load_past_appointments(self):
        conn = create_connection()
        c = conn.cursor()
        c.execute("SELECT adsoyad, pasaport, randevu_tarihi, randevu_saati, ofis FROM randevular")
        rows = c.fetchall()
        conn.close()

        self.past_appointments_table.setRowCount(len(rows))
        for row_index, row_data in enumerate(rows):
            for col_index, col_data in enumerate(row_data):
               item = QTableWidgetItem(str(col_data))
               self.past_appointments_table.setItem(row_index, col_index, item)
    
    
dark_style = """
    QWidget {
        background-color: #001f3f;  /* Lacivert zemin */
        color: white;
        font-family: 'Segoe UI', Arial;
        font-size: 14px;
    }

    QLabel#header {
        font-size: 20px;
        font-weight: bold;
        color: #66b2ff;  /* Parlak mavi başlık */
        margin-bottom: 10px;
    }

    QPushButton {
        background-color: #004080;
        color: white;
        border: 1px solid #0066cc;
        padding: 8px 14px;
        border-radius: 6px;
    }

    QPushButton:hover {
        background-color: #0059b3;
        border: 1px solid #3399ff;
    }

    QLineEdit, QTextEdit, QComboBox, QDateEdit, QTimeEdit {
        background-color: #002b4d;
        color: white;
        border: 1px solid #3399ff;
        border-radius: 4px;
        padding: 4px;
    }

    QTabWidget::pane {
        border: 1px solid #003366;
        background: #001f3f;
    }

    QTabBar::tab {
        background: #002b4d;
        color: white;
        padding: 8px 12px;
        border-top-left-radius: 6px;
        border-top-right-radius: 6px;
        margin-right: 2px;
    }

    QTabBar::tab:selected {
        background: #004080;
        font-weight: bold;
        color: #66ccff;
    }

    QGroupBox {
        border: 1px solid #005580;
        border-radius: 5px;
        margin-top: 10px;
    }

    QGroupBox:title {
        subcontrol-origin: margin;
        subcontrol-position: top left;
        padding: 0 5px;
        color: #66ccff;
    }
    QPushButton#greenButton {
    background-color: #28a745;  /* Bootstrap yeşili */
    color: white;
    border: 1px solid #218838;
    padding: 6px;
    border-radius: 4px;
}

QPushButton#greenButton:hover {
    background-color: #218838;
}

"""


app = QApplication(sys.argv)
app.setStyleSheet(dark_style)

window = RandevuUygulamasi()
window.show()
sys.exit(app.exec_())
